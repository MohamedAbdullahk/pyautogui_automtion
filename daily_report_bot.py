"""
Gen AI Architect Program - Assignment 1: PyAutoGUI Automation
Daily Status Report Bot

What this bot does (matches the assignment steps exactly):
1. Opens Chrome and goes to a public news website (The Hindu).
2. Reads the top headline from the page (via HTML, so it's always the
   real headline and never an ad or promo banner - see Step 3 below).
3. Opens a spreadsheet app (WPS Spreadsheet instead of Excel, since Excel
   is blocked on this laptop - WPS reads/writes the same .xlsx format).
4. Adds a new row with: today's date & time, the headline, and a comment.
5. Saves the file with today's date in the filename.
6. Takes a screenshot of the final sheet and saves it.

Note: PyAutoGUI controls the real mouse/keyboard, so screen coordinates
below (the numbers in pyautogui.click(x, y)) may need small adjustments
to match YOUR screen resolution. Run once, watch what happens, and tweak.

One-time setup on your laptop:
    pip install pyautogui requests beautifulsoup4 openpyxl
"""

import os
import subprocess
import time
from datetime import datetime

import pyautogui
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook   # helper library used to read/write the .xlsx file

# Slow down a little so Windows/apps have time to react to each action
pyautogui.PAUSE = 1

# Give ourselves a few seconds after starting the script to switch focus
# to the screen (in case you're clicking "Run" from an editor)
time.sleep(3)


# -----------------------------------------------------------------
# STEP 1 & 2: Open Chrome at a fixed window size/position, already
# pointed at The Hindu.
#
# Instead of pressing Win -> typing "chrome" -> waiting for search
# results (timing can vary), we launch Chrome directly with command
# line flags. This also fixes the window to the same size and screen
# position every time, which matters for PyAutoGUI's other clicks.
# -----------------------------------------------------------------
CHROME_PATHS = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
]
chrome_exe = next((p for p in CHROME_PATHS if os.path.exists(p)), None)

if chrome_exe:
    subprocess.Popen([
        chrome_exe,
        "--new-window",
        "--window-size=1280,800",   # width,height in pixels
        "--window-position=0,0",    # top-left corner of the screen
        "https://www.thehindu.com",
    ])
else:
    # Fallback: Chrome not found at the usual paths - open it the
    # old way via the Start menu search instead.
    pyautogui.press('win')
    time.sleep(1)
    pyautogui.write('chrome')
    pyautogui.press('enter')
    time.sleep(5)
    pyautogui.hotkey('win', 'up')   # maximize as a best-effort fallback
    time.sleep(1)
    pyautogui.write('https://www.thehindu.com')
    pyautogui.press('enter')

time.sleep(6)  # wait for the page to load


# -----------------------------------------------------------------
# STEP 3: Get the top headline from the page
# -----------------------------------------------------------------
# WHY NOT click+drag+copy on screen? Because a fixed pixel position like
# (500, 400) can land on an ad banner, a promo carousel, or a cookie
# pop-up instead of the actual headline - the homepage layout shifts
# between visits. To reliably get the REAL top story, we read the page's
# HTML directly (same page Chrome is showing) instead of guessing coordinates.
def get_top_headline():
    try:
        response = requests.get(
            "https://www.thehindu.com",
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10,
        )
        soup = BeautifulSoup(response.text, "html.parser")

        # News sites mark their main headline with schema.org "headline"
        # metadata for SEO - this is the most reliable way to find it.
        tag = soup.find(attrs={"itemprop": "headline"})
        if not tag:
            tag = soup.find("h1")          # fallback: first <h1> on the page
        if not tag:
            tag = soup.find("h3")          # fallback: first <h3> on the page

        text = tag.get_text(strip=True) if tag else ""
        return text if text else None
    except Exception as e:
        print(f"Could not fetch headline from HTML: {e}")
        return None


headline = get_top_headline()
if not headline:
    headline = "Could not read headline - check internet/site access"

print(f"Top headline found: {headline}")


# -----------------------------------------------------------------
# STEP 4: Create today's spreadsheet (if it doesn't exist) and open it
#          in WPS Spreadsheet
# -----------------------------------------------------------------
today = datetime.now().strftime("%Y-%m-%d")
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          f"daily_report_{today}.xlsx")

if not os.path.exists(file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Report"
    sheet.append(["Date & Time", "Headline", "Comment"])  # header row (row 1)
    workbook.save(file_path)

# Figure out which row the new entry should go in by reading the file
# directly, instead of guessing on screen with Ctrl+End (which Excel/WPS
# can report incorrectly if old formatting or cached ranges are present).
existing = load_workbook(file_path)
next_row = existing.active.max_row + 1   # e.g. max_row=1 (just header) -> row 2
existing.close()

os.startfile(file_path)   # opens the file in WPS Spreadsheet
time.sleep(8)              # wait for WPS to fully open


# -----------------------------------------------------------------
# STEP 5: Click into the sheet, then move down to the exact next
#          empty row using plain arrow-key navigation.
#
# WHY NOT Ctrl+G (Go To)? It relies on a dialog box opening and
# grabbing keyboard focus - if it opens a moment too slowly (or WPS
# handles the shortcut slightly differently than Excel), the typed
# cell reference lands directly in whatever cell is selected instead
# of in the dialog, which overwrites real data (e.g. the header row).
#
# Arrow-key movement has no such risk: when the file opens, the
# cursor always starts at A1, so pressing Down a known number of
# times always lands exactly on the row we want.
# -----------------------------------------------------------------
pyautogui.click(300, 300)        # click somewhere inside the spreadsheet grid
time.sleep(1)

pyautogui.hotkey('ctrl', 'home')  # make sure we start from cell A1
time.sleep(1)

for _ in range(next_row - 2):     # e.g. next_row=2 -> press Down once
    pyautogui.press('down')
pyautogui.press('home')           # make sure we're in column A of that row
time.sleep(1)


# -----------------------------------------------------------------
# STEP 6: Type the new row - date/time, headline, comment
# -----------------------------------------------------------------
now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")   # generated at run time
comment = "Top news captured automatically"

pyautogui.write(now)
pyautogui.press('tab')
pyautogui.write(headline)
pyautogui.press('tab')
pyautogui.write(comment)
pyautogui.press('enter')


# -----------------------------------------------------------------
# STEP 7: Save the spreadsheet
# -----------------------------------------------------------------
time.sleep(1)
pyautogui.hotkey('ctrl', 's')
time.sleep(2)
pyautogui.press('enter')   # confirms "Keep current format" if WPS asks
time.sleep(2)


# -----------------------------------------------------------------
# STEP 8: Take a screenshot of the final sheet
# -----------------------------------------------------------------
screenshot_path = os.path.join(os.path.dirname(file_path),
                                f"daily_report_{today}.png")
pyautogui.screenshot().save(screenshot_path)

print("Done!")
print(f"Spreadsheet saved to : {file_path}")
print(f"Screenshot saved to  : {screenshot_path}")